#!/usr/bin/env python3
"""
SAC Template Generator v2
Universal tool for generating Excel templates for SAP Analytics Cloud mass uploads.

Usage:
    python sac_generator.py --init              # Create empty configuration
    python sac_generator.py --download          # Download all dimension CSVs from SAC
    python sac_generator.py --scan              # Scan CSVs and suggest dimensions config
    python sac_generator.py --list              # List available templates
    python sac_generator.py --validate          # Validate configuration
    python sac_generator.py --generate          # Generate all templates
    python sac_generator.py --generate "Name"   # Generate specific template
"""

import argparse
import json
import os
import sys
import re
import logging
import traceback
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from enum import Enum
from dataclasses import dataclass, field

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# Try to import browser_cookie3 for automatic cookie extraction
try:
    import browser_cookie3
    BROWSER_COOKIES_AVAILABLE = True
except ImportError:
    BROWSER_COOKIES_AVAILABLE = False


# =============================================================================
# LOGGING SETUP
# =============================================================================

def setup_logging(log_dir: str = "logs", level: str = "INFO") -> logging.Logger:
    """Setup logging with file and console handlers."""
    os.makedirs(log_dir, exist_ok=True)
    
    logger = logging.getLogger("sac_generator")
    logger.setLevel(getattr(logging, level.upper(), logging.INFO))
    
    # Clear existing handlers
    logger.handlers = []
    
    # File handler - detailed JSON logs
    log_file = os.path.join(log_dir, f"sac_generator_{datetime.now().strftime('%Y%m%d')}.log")
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter(
        '{"timestamp": "%(asctime)s", "level": "%(levelname)s", "module": "%(module)s", '
        '"function": "%(funcName)s", "line": %(lineno)d, "message": "%(message)s"}'
    )
    file_handler.setFormatter(file_formatter)
    logger.addHandler(file_handler)
    
    # Console handler - human readable
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_formatter = logging.Formatter('%(levelname)s: %(message)s')
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)
    
    return logger

# Global logger
logger = setup_logging()


# =============================================================================
# CUSTOM EXCEPTIONS
# =============================================================================

class SACGeneratorError(Exception):
    """Base exception for SAC Generator."""
    def __init__(self, message: str, details: dict = None):
        self.message = message
        self.details = details or {}
        super().__init__(self.message)
    
    def to_dict(self) -> dict:
        return {
            "error_type": self.__class__.__name__,
            "message": self.message,
            "details": self.details
        }


class ConfigurationError(SACGeneratorError):
    """Raised when configuration is invalid or missing."""
    pass


class CSVError(SACGeneratorError):
    """Raised when CSV file operations fail."""
    pass


class DimensionError(SACGeneratorError):
    """Raised when dimension processing fails."""
    pass


class TemplateError(SACGeneratorError):
    """Raised when template operations fail."""
    pass


class SACConnectionError(SACGeneratorError):
    """Raised when SAC connection fails."""
    pass


class ExcelGenerationError(SACGeneratorError):
    """Raised when Excel generation fails."""
    pass


class ValidationError(SACGeneratorError):
    """Raised when validation fails."""
    pass


# =============================================================================
# RESULT TYPES
# =============================================================================

class ResultStatus(Enum):
    SUCCESS = "success"
    WARNING = "warning"
    ERROR = "error"


@dataclass
class OperationResult:
    """Result of an operation with status, message, and data."""
    status: ResultStatus
    message: str
    data: Any = None
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    
    @property
    def success(self) -> bool:
        return self.status in (ResultStatus.SUCCESS, ResultStatus.WARNING)
    
    def to_dict(self) -> dict:
        return {
            "status": self.status.value,
            "success": self.success,
            "message": self.message,
            "data": self.data,
            "errors": self.errors,
            "warnings": self.warnings
        }
    
    @classmethod
    def ok(cls, message: str = "OK", data: Any = None, warnings: List[str] = None):
        return cls(
            status=ResultStatus.WARNING if warnings else ResultStatus.SUCCESS,
            message=message,
            data=data,
            warnings=warnings or []
        )
    
    @classmethod
    def fail(cls, message: str, errors: List[str] = None):
        return cls(
            status=ResultStatus.ERROR,
            message=message,
            errors=errors or [message]
        )


class SACGenerator:
    """Main class for SAC Template Generator."""
    
    def __init__(self, config_dir: str = "config"):
        self.config_dir = config_dir
        self.downloads_dir = "downloads"
        self.output_dir = "output"
        self.errors: List[str] = []
        self.warnings: List[str] = []
        
        logger.info(f"Initializing SACGenerator with config_dir={config_dir}")
        
        # Ensure directories exist
        try:
            os.makedirs(self.config_dir, exist_ok=True)
            os.makedirs(self.downloads_dir, exist_ok=True)
            os.makedirs(self.output_dir, exist_ok=True)
        except OSError as e:
            logger.error(f"Failed to create directories: {e}")
            raise ConfigurationError(
                f"Cannot create required directories",
                {"config_dir": self.config_dir, "error": str(e)}
            )
        
        # Load configurations with error handling
        self.project = self._load_json("project.json", {})
        self.dimensions_config = self._load_json("dimensions.json", {"dimensions": []})
        self.templates_config = self._load_json("templates.json", {"templates": []})
        
        # Runtime data
        self.dimension_data = {}
        self.date_range = []
        
        logger.info("SACGenerator initialized successfully")
    
    def _load_json(self, filename: str, default: dict) -> dict:
        """Load JSON configuration file with error handling."""
        filepath = os.path.join(self.config_dir, filename)
        
        if not os.path.exists(filepath):
            logger.debug(f"Config file not found, using default: {filepath}")
            return default
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                logger.debug(f"Loaded config: {filepath}")
                return data
        except json.JSONDecodeError as e:
            error_msg = f"Invalid JSON in {filename}: {e.msg} at line {e.lineno}"
            logger.error(error_msg)
            raise ConfigurationError(
                error_msg,
                {"file": filename, "line": e.lineno, "column": e.colno}
            )
        except IOError as e:
            error_msg = f"Cannot read {filename}: {str(e)}"
            logger.error(error_msg)
            raise ConfigurationError(error_msg, {"file": filename})
    
    def _save_json(self, filename: str, data: dict) -> OperationResult:
        """Save JSON configuration file with error handling."""
        filepath = os.path.join(self.config_dir, filename)
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            logger.info(f"Saved config: {filepath}")
            print(f"  Saved: {filepath}")
            return OperationResult.ok(f"Saved {filename}")
        except IOError as e:
            error_msg = f"Cannot write {filename}: {str(e)}"
            logger.error(error_msg)
            raise ConfigurationError(error_msg, {"file": filename})
        except TypeError as e:
            error_msg = f"Cannot serialize data to JSON: {str(e)}"
            logger.error(error_msg)
            raise ConfigurationError(error_msg, {"file": filename})
    
    # =========================================================================
    # INITIALIZATION
    # =========================================================================
    
    def init_config(self):
        """Create empty configuration files with examples."""
        print("\n=== Initializing SAC Template Generator ===\n")
        
        # Project config
        project = {
            "name": "My SAC Project",
            "description": "SAC Template Generator Project",
            "sac_connection": {
                "base_url": "https://YOUR-TENANT.eu20.analytics.cloud.sap",
                "model_id": "YOUR_MODEL_ID",
                "version_model_id": "YOUR_VERSION_MODEL_ID"
            },
            "version": {
                "version_id": "public.RF_CURRENT"
            },
            "settings": {
                "data_rows": 200,
                "remove_unassigned": True,
                "remove_not_in_hierarchy": True
            }
        }
        self._save_json("project.json", project)
        
        # Dimensions config
        dimensions = {
            "_comment": "Run 'python sac_generator.py --scan' to auto-detect dimensions",
            "dimensions": []
        }
        self._save_json("dimensions.json", dimensions)
        
        # Templates config
        templates = {
            "_comment": "Define your templates here",
            "templates": []
        }
        self._save_json("templates.json", templates)
        
        print("\n✓ Configuration initialized!")
        print("\nNext steps:")
        print("  1. Edit config/project.json with your SAC connection details")
        print("  2. Run: python sac_generator.py --download")
        print("  3. Run: python sac_generator.py --scan")
        print("  4. Edit config/templates.json to define your templates")
        print("  5. Run: python sac_generator.py --generate")
    
    # =========================================================================
    # DOWNLOAD DIMENSIONS
    # =========================================================================
    
    def get_browser_cookies(self) -> Optional[requests.cookies.RequestsCookieJar]:
        """Extract cookies from browser for SAC authentication."""
        logger.info("Attempting to load browser cookies")
        
        if not BROWSER_COOKIES_AVAILABLE:
            logger.warning("browser_cookie3 not installed")
            print("  WARNING: browser_cookie3 not installed")
            print("  Install with: pip install browser_cookie3")
            return None
        
        base_url = self.project.get("sac_connection", {}).get("base_url", "")
        if not base_url:
            logger.error("SAC base_url not configured")
            raise ConfigurationError(
                "SAC base_url not configured in project.json",
                {"required_field": "sac_connection.base_url"}
            )
        
        domain = base_url.replace("https://", "").replace("http://", "").split("/")[0]
        logger.debug(f"Looking for cookies for domain: {domain}")
        
        # Try different browsers
        browsers = [
            ("Chrome", browser_cookie3.chrome),
            ("Edge", browser_cookie3.edge),
            ("Firefox", browser_cookie3.firefox),
        ]
        
        browser_errors = []
        for browser_name, browser_func in browsers:
            try:
                cookies = browser_func(domain_name=domain)
                logger.info(f"Successfully loaded cookies from {browser_name}")
                print(f"  ✓ Loaded cookies from {browser_name}")
                return cookies
            except Exception as e:
                error_detail = f"{browser_name}: {str(e)}"
                logger.debug(f"Failed to load cookies from {browser_name}: {e}")
                browser_errors.append(error_detail)
                continue
        
        logger.warning(f"Could not load cookies from any browser: {browser_errors}")
        print("  WARNING: Could not load cookies from any browser")
        print("  Make sure you are logged into SAC in your browser")
        return None
    
    def build_download_url(self, sac_name: str, has_hierarchy: bool) -> str:
        """Build download URL for a dimension."""
        base_url = self.project.get("sac_connection", {}).get("base_url", "")
        model_id = self.project.get("sac_connection", {}).get("model_id", "")
        
        if not base_url or not model_id:
            raise ConfigurationError(
                "SAC connection not fully configured",
                {"base_url": bool(base_url), "model_id": bool(model_id)}
            )
        
        suffix = "MasterWithHierarchy" if has_hierarchy else "Master"
        url = f"{base_url}/api/v1/dataexport/providers/sac/{model_id}/{sac_name}{suffix}?$format=text/csv"
        logger.debug(f"Built download URL: {url}")
        return url
    
    def download_dimension(self, sac_name: str, has_hierarchy: bool, cookies) -> OperationResult:
        """Download a single dimension CSV."""
        url = self.build_download_url(sac_name, has_hierarchy)
        suffix = "MasterWithHierarchy" if has_hierarchy else "Master"
        filename = f"{sac_name}{suffix}.csv"
        filepath = os.path.join(self.downloads_dir, filename)
        
        logger.info(f"Downloading dimension: {sac_name} (hierarchy={has_hierarchy})")
        print(f"  Downloading {sac_name}...")
        
        try:
            response = requests.get(url, cookies=cookies, timeout=60)
            
            if response.status_code == 200:
                # Validate response content
                content = response.content
                if len(content) < 10:
                    logger.warning(f"Downloaded file {filename} is suspiciously small ({len(content)} bytes)")
                    return OperationResult.ok(
                        f"Downloaded {filename} (warning: only {len(content)} bytes)",
                        warnings=[f"File is very small ({len(content)} bytes)"]
                    )
                
                with open(filepath, 'wb') as f:
                    f.write(content)
                
                logger.info(f"Saved: {filename} ({len(content)} bytes)")
                print(f"    ✓ Saved: {filename}")
                return OperationResult.ok(f"Downloaded {filename}", data={"path": filepath, "size": len(content)})
            
            elif response.status_code == 401:
                error_msg = f"Authentication failed for {sac_name} - cookies may have expired"
                logger.error(error_msg)
                print(f"    ✗ Authentication failed - please re-login to SAC in your browser")
                return OperationResult.fail(error_msg)
            
            elif response.status_code == 403:
                error_msg = f"Access denied for {sac_name} - check your SAC permissions"
                logger.error(error_msg)
                print(f"    ✗ Access denied - check your permissions")
                return OperationResult.fail(error_msg)
            
            elif response.status_code == 404:
                error_msg = f"Dimension {sac_name} not found - check dimension name and model ID"
                logger.error(error_msg)
                print(f"    ✗ Not found - check dimension name")
                return OperationResult.fail(error_msg)
            
            else:
                error_msg = f"HTTP {response.status_code}: {response.reason}"
                logger.error(f"Download failed for {sac_name}: {error_msg}")
                print(f"    ✗ Error {response.status_code}: {response.reason}")
                return OperationResult.fail(error_msg)
                
        except requests.exceptions.Timeout:
            error_msg = f"Timeout downloading {sac_name} - SAC may be slow or unreachable"
            logger.error(error_msg)
            print(f"    ✗ Timeout - SAC may be slow")
            return OperationResult.fail(error_msg)
        
        except requests.exceptions.ConnectionError as e:
            error_msg = f"Connection error for {sac_name} - check network and SAC URL"
            logger.error(f"{error_msg}: {e}")
            print(f"    ✗ Connection error - check network")
            return OperationResult.fail(error_msg)
        
        except Exception as e:
            error_msg = f"Unexpected error downloading {sac_name}: {str(e)}"
            logger.exception(error_msg)
            print(f"    ✗ Error: {str(e)}")
            return OperationResult.fail(error_msg)
    
    def download_version(self, cookies) -> OperationResult:
        """Download Version dimension for date range."""
        base_url = self.project.get("sac_connection", {}).get("base_url", "")
        version_model_id = self.project.get("sac_connection", {}).get("version_model_id", "")
        
        if not version_model_id:
            logger.warning("version_model_id not configured, skipping Version download")
            return OperationResult.ok("Skipped Version download (not configured)", warnings=["version_model_id not set"])
        
        url = f"{base_url}/api/v1/dataexport/providers/sac/{version_model_id}/VersionMaster?$format=text/csv"
        filepath = os.path.join(self.downloads_dir, "VersionMaster.csv")
        
        logger.info("Downloading Version dimension")
        print(f"  Downloading Version...")
        
        try:
            response = requests.get(url, cookies=cookies, timeout=60)
            
            if response.status_code == 200:
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                logger.info(f"Saved VersionMaster.csv ({len(response.content)} bytes)")
                print(f"    ✓ Saved: VersionMaster.csv")
                return OperationResult.ok("Downloaded VersionMaster.csv")
            else:
                error_msg = f"HTTP {response.status_code} downloading Version"
                logger.error(error_msg)
                print(f"    ✗ Error {response.status_code}")
                return OperationResult.fail(error_msg)
                
        except Exception as e:
            error_msg = f"Error downloading Version: {str(e)}"
            logger.exception(error_msg)
            print(f"    ✗ Error: {str(e)}")
            return OperationResult.fail(error_msg)
    
    def download_all(self) -> OperationResult:
        """Download all configured dimensions."""
        logger.info("Starting download of all dimensions")
        print("\n=== Downloading Dimensions from SAC ===\n")
        
        results = {"success": [], "failed": [], "warnings": []}
        
        # Check configuration
        base_url = self.project.get("sac_connection", {}).get("base_url", "")
        if "YOUR-TENANT" in base_url or not base_url:
            error_msg = "SAC connection not configured in project.json"
            logger.error(error_msg)
            print("ERROR: Please configure config/project.json first")
            print("       Set your SAC base_url and model_id")
            return OperationResult.fail(error_msg)
        
        # Get cookies
        try:
            cookies = self.get_browser_cookies()
        except SACGeneratorError as e:
            return OperationResult.fail(str(e))
        
        if cookies is None:
            error_msg = "Cannot proceed without browser cookies"
            logger.error(error_msg)
            print("\nERROR: Cannot proceed without browser cookies")
            print("Make sure you are logged into SAC in your browser")
            return OperationResult.fail(error_msg)
        
        # Collect unique SAC dimensions to download
        dimensions_to_download = {}
        for dim in self.dimensions_config.get("dimensions", []):
            sac_name = dim.get("sac_name")
            if not sac_name:
                results["warnings"].append(f"Dimension '{dim.get('name', 'unknown')}' has no sac_name")
                continue
            has_hierarchy = dim.get("has_hierarchy", True)
            key = f"{sac_name}_{has_hierarchy}"
            if key not in dimensions_to_download:
                dimensions_to_download[key] = (sac_name, has_hierarchy)
        
        if not dimensions_to_download:
            logger.warning("No dimensions configured to download")
            print("No dimensions configured in config/dimensions.json")
            print("Run --scan after manually downloading some CSVs, or add dimensions manually")
            return OperationResult.fail("No dimensions configured")
        
        # Download each dimension
        for sac_name, has_hierarchy in dimensions_to_download.values():
            result = self.download_dimension(sac_name, has_hierarchy, cookies)
            if result.success:
                results["success"].append(sac_name)
                results["warnings"].extend(result.warnings)
            else:
                results["failed"].append({"dimension": sac_name, "error": result.message})
        
        # Download Version
        version_result = self.download_version(cookies)
        results["warnings"].extend(version_result.warnings)
        
        success_count = len(results["success"])
        failed_count = len(results["failed"])
        
        logger.info(f"Download complete: {success_count} success, {failed_count} failed")
        print(f"\n✓ Downloaded: {success_count}, Failed: {failed_count}")
        
        if failed_count > 0:
            return OperationResult(
                status=ResultStatus.WARNING,
                message=f"Downloaded {success_count}, failed {failed_count}",
                data=results,
                warnings=results["warnings"],
                errors=[f["error"] for f in results["failed"]]
            )
        
        return OperationResult.ok(
            f"Downloaded all {success_count} dimensions",
            data=results,
            warnings=results["warnings"]
        )
    
    # =========================================================================
    # SCAN CSV FILES
    # =========================================================================
    
    def scan_csv_files(self):
        """Scan downloads folder and suggest dimensions configuration."""
        print("\n=== Scanning CSV Files ===\n")
        
        csv_files = [f for f in os.listdir(self.downloads_dir) if f.endswith('.csv')]
        
        if not csv_files:
            print("No CSV files found in downloads/")
            print("Either download manually or run --download first")
            return
        
        print(f"Found {len(csv_files)} CSV files:\n")
        
        suggested_dimensions = []
        
        for csv_file in sorted(csv_files):
            filepath = os.path.join(self.downloads_dir, csv_file)
            
            # Parse filename
            has_hierarchy = "WithHierarchy" in csv_file
            
            # Extract SAC name
            sac_name = csv_file.replace("MasterWithHierarchy.csv", "").replace("Master.csv", "").replace(".csv", "")
            
            # Skip Version
            if sac_name == "Version":
                print(f"  {csv_file} -> Version (for date range)")
                continue
            
            # Load and analyze CSV
            try:
                df = pd.read_csv(filepath, dtype=str, nrows=100)
                columns = list(df.columns)
                row_count = len(pd.read_csv(filepath, dtype=str))
                
                # Detect hierarchy columns
                parent_cols = [c for c in columns if c.endswith('_PARENTID')]
                
                # Create dimension suggestion
                dim_name = sac_name.replace("COL_", "").replace("_", " ").title()
                table_name = f"tbl_{sac_name.lower().replace('col_', '')}"
                
                dim = {
                    "name": dim_name,
                    "sac_name": sac_name,
                    "has_hierarchy": has_hierarchy,
                    "table_name": table_name
                }
                
                suggested_dimensions.append(dim)
                
                print(f"  {csv_file}")
                print(f"    -> Name: {dim_name}")
                print(f"    -> Columns: {', '.join(columns[:5])}{'...' if len(columns) > 5 else ''}")
                print(f"    -> Rows: {row_count}, Hierarchy: {has_hierarchy}")
                if parent_cols:
                    print(f"    -> Parent columns: {', '.join(parent_cols)}")
                print()
                
            except Exception as e:
                print(f"  {csv_file} -> Error: {str(e)}")
        
        # Ask to save
        print("\n" + "="*50)
        response = input("\nSave suggested dimensions to config/dimensions.json? [y/N]: ")
        
        if response.lower() == 'y':
            self.dimensions_config["dimensions"] = suggested_dimensions
            self._save_json("dimensions.json", self.dimensions_config)
            print("\n✓ Dimensions saved!")
            print("  Edit config/dimensions.json to add filters, rename dimensions, etc.")
    
    # =========================================================================
    # LIST & VALIDATE
    # =========================================================================
    
    def list_templates(self):
        """List all configured templates."""
        print("\n=== Configured Templates ===\n")
        
        templates = self.templates_config.get("templates", [])
        
        if not templates:
            print("No templates configured in config/templates.json")
            return
        
        for i, template in enumerate(templates, 1):
            name = template.get("name", "Unnamed")
            output = template.get("output_file", "")
            columns = template.get("columns", [])
            
            print(f"{i}. {name}")
            print(f"   Output: {output}")
            print(f"   Columns: {', '.join(columns)}")
            print()
    
    def validate_config(self):
        """Validate configuration files."""
        print("\n=== Validating Configuration ===\n")
        
        errors = []
        warnings = []
        
        # Check project config
        base_url = self.project.get("sac_connection", {}).get("base_url", "")
        if "YOUR-TENANT" in base_url or not base_url:
            errors.append("project.json: base_url not configured")
        
        model_id = self.project.get("sac_connection", {}).get("model_id", "")
        if "YOUR_MODEL_ID" in model_id or not model_id:
            errors.append("project.json: model_id not configured")
        
        # Check dimensions
        dimensions = {d["name"]: d for d in self.dimensions_config.get("dimensions", [])}
        if not dimensions:
            warnings.append("dimensions.json: No dimensions defined")
        
        # Check templates
        templates = self.templates_config.get("templates", [])
        if not templates:
            warnings.append("templates.json: No templates defined")
        
        for template in templates:
            name = template.get("name", "Unnamed")
            for col in template.get("columns", []):
                if col not in dimensions:
                    errors.append(f"Template '{name}': Dimension '{col}' not found in dimensions.json")
        
        # Check CSV files
        for dim in dimensions.values():
            sac_name = dim.get("sac_name")
            has_hierarchy = dim.get("has_hierarchy", True)
            suffix = "MasterWithHierarchy" if has_hierarchy else "Master"
            csv_file = f"{sac_name}{suffix}.csv"
            csv_path = os.path.join(self.downloads_dir, csv_file)
            
            if not os.path.exists(csv_path):
                warnings.append(f"CSV not found: {csv_file} (for dimension '{dim['name']}')")
        
        # Report
        if errors:
            print("ERRORS:")
            for e in errors:
                print(f"  ✗ {e}")
            print()
        
        if warnings:
            print("WARNINGS:")
            for w in warnings:
                print(f"  ⚠ {w}")
            print()
        
        if not errors and not warnings:
            print("✓ All configuration valid!")
        elif not errors:
            print("✓ Configuration valid (with warnings)")
        else:
            print("✗ Configuration has errors - please fix before generating")
    
    # =========================================================================
    # GENERATE TEMPLATES
    # =========================================================================
    
    def find_csv_file(self, sac_name: str, has_hierarchy: bool) -> Optional[str]:
        """Find CSV file for a dimension."""
        if has_hierarchy:
            patterns = [f"{sac_name}MasterWithHierarchy.csv", f"{sac_name}Master.csv", f"{sac_name}.csv"]
        else:
            patterns = [f"{sac_name}Master.csv", f"{sac_name}.csv", f"{sac_name}MasterWithHierarchy.csv"]
        
        for pattern in patterns:
            path = os.path.join(self.downloads_dir, pattern)
            if os.path.exists(path):
                logger.debug(f"Found CSV file: {path}")
                return path
        
        logger.warning(f"No CSV file found for {sac_name}")
        return None
    
    def load_csv(self, csv_path: str) -> Optional[pd.DataFrame]:
        """Load CSV file with comprehensive error handling."""
        if not os.path.exists(csv_path):
            logger.error(f"CSV file not found: {csv_path}")
            return None
        
        try:
            df = pd.read_csv(csv_path, dtype=str, quotechar='"', escapechar='\\')
            logger.debug(f"Loaded CSV: {csv_path} ({len(df)} rows, {len(df.columns)} cols)")
            return df
        
        except pd.errors.EmptyDataError:
            error_msg = f"CSV file is empty: {csv_path}"
            logger.error(error_msg)
            raise CSVError(error_msg, {"path": csv_path})
        
        except pd.errors.ParserError as e:
            error_msg = f"CSV parsing error in {csv_path}: {str(e)}"
            logger.error(error_msg)
            raise CSVError(error_msg, {"path": csv_path, "parser_error": str(e)})
        
        except UnicodeDecodeError as e:
            # Try with different encoding
            logger.warning(f"UTF-8 decode failed for {csv_path}, trying latin-1")
            try:
                df = pd.read_csv(csv_path, dtype=str, encoding='latin-1', quotechar='"', escapechar='\\')
                logger.info(f"Loaded CSV with latin-1 encoding: {csv_path}")
                return df
            except Exception as e2:
                error_msg = f"Cannot decode CSV {csv_path}: {str(e)}"
                logger.error(error_msg)
                raise CSVError(error_msg, {"path": csv_path, "encoding_error": str(e)})
        
        except Exception as e:
            error_msg = f"Unexpected error loading {csv_path}: {str(e)}"
            logger.exception(error_msg)
            raise CSVError(error_msg, {"path": csv_path})
    
    def find_parent_columns(self, df: pd.DataFrame) -> List[str]:
        """Find hierarchy parent columns."""
        return [col for col in df.columns if col.endswith('_PARENTID')]
    
    def get_leaf_members(self, df: pd.DataFrame, parent_filter=None) -> pd.DataFrame:
        """Filter to leaf members only."""
        if df is None or df.empty:
            return df
        
        parent_cols = self.find_parent_columns(df)
        
        if not parent_cols:
            return df
        
        # Filter to parent subtree if specified
        if parent_filter:
            if isinstance(parent_filter, str):
                parent_filter = [parent_filter]
            
            descendants = set()
            to_process = set(parent_filter)
            
            while to_process:
                current = to_process.pop()
                for col in parent_cols:
                    children = df[df[col] == current]['ID'].tolist()
                    for child in children:
                        if child not in descendants:
                            descendants.add(child)
                            to_process.add(child)
            
            df = df[df['ID'].isin(descendants)].copy()
        
        # Get leaf members
        all_parent_ids = set()
        for col in parent_cols:
            all_parent_ids.update(df[col].dropna().unique())
        
        leaf_mask = ~df['ID'].isin(all_parent_ids)
        return df[leaf_mask].copy()
    
    def remove_unassigned(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove unassigned and 'Not In Hierarchy' members."""
        if df is None or df.empty:
            return df
        
        mask = ~df['ID'].isin(['#', ''])
        if 'Description' in df.columns:
            mask &= ~df['Description'].isin(['Unassigned', 'Not In Hierarchy', ''])
        
        return df[mask].copy()
    
    def apply_filters(self, df: pd.DataFrame, filters: dict) -> pd.DataFrame:
        """Apply dimension filters."""
        if df is None or df.empty or not filters:
            return df
        
        # Exclude by description
        exclude_desc = filters.get("exclude_description", [])
        if exclude_desc and 'Description' in df.columns:
            for pattern in exclude_desc:
                mask = ~df['Description'].str.contains(pattern, case=False, na=False)
                df = df[mask]
        
        # Filter by ID list
        id_list = filters.get("id_list", [])
        if id_list:
            df = df[df['ID'].isin(id_list)].copy()
            # Sort by id_list order
            df['_sort'] = df['ID'].apply(lambda x: id_list.index(x) if x in id_list else 999)
            df = df.sort_values('_sort').drop('_sort', axis=1)
        
        return df
    
    def extract_column_values(self, sac_name: str, column: str, has_hierarchy: bool, numeric_sort: bool = False) -> pd.DataFrame:
        """Extract unique values from a column."""
        csv_path = self.find_csv_file(sac_name, has_hierarchy)
        if not csv_path:
            return pd.DataFrame(columns=['ID', 'Description'])
        
        df = self.load_csv(csv_path)
        if df is None or column not in df.columns:
            return pd.DataFrame(columns=['ID', 'Description'])
        
        values = df[column].dropna().unique()
        values = [v for v in values if v and str(v).strip() and v not in ['#', 'Not In Hierarchy']]
        
        if numeric_sort:
            try:
                values = sorted(values, key=lambda x: int(float(x)))
            except:
                values = sorted(values)
        else:
            values = sorted(values)
        
        return pd.DataFrame({'ID': values, 'Description': values})
    
    def load_dimension(self, dim_config: dict) -> pd.DataFrame:
        """Load and process a dimension with comprehensive error handling."""
        dim_name = dim_config.get("name", "unknown")
        sac_name = dim_config.get("sac_name")
        has_hierarchy = dim_config.get("has_hierarchy", True)
        filters = dim_config.get("filters", {})
        extract_column = dim_config.get("extract_column")
        numeric_sort = dim_config.get("numeric_sort", False)
        
        logger.info(f"Loading dimension: {dim_name} (sac_name={sac_name})")
        
        if not sac_name:
            error_msg = f"Dimension '{dim_name}' has no sac_name configured"
            logger.error(error_msg)
            raise DimensionError(error_msg, {"dimension": dim_name})
        
        # Special case: extract column values
        if extract_column:
            logger.debug(f"Extracting column values: {extract_column}")
            try:
                result = self.extract_column_values(sac_name, extract_column, has_hierarchy, numeric_sort)
                logger.info(f"Extracted {len(result)} values from column {extract_column}")
                return result
            except Exception as e:
                logger.exception(f"Error extracting column values for {dim_name}")
                raise DimensionError(
                    f"Failed to extract column {extract_column} for {dim_name}: {str(e)}",
                    {"dimension": dim_name, "column": extract_column}
                )
        
        csv_path = self.find_csv_file(sac_name, has_hierarchy)
        if not csv_path:
            warning_msg = f"CSV file not found for {sac_name}"
            logger.warning(warning_msg)
            self.warnings.append(warning_msg)
            print(f"  WARNING: CSV file not found for {sac_name}")
            return pd.DataFrame(columns=['ID', 'Description'])
        
        try:
            df = self.load_csv(csv_path)
        except CSVError as e:
            logger.error(f"Failed to load CSV for {dim_name}: {e}")
            self.errors.append(str(e))
            return pd.DataFrame(columns=['ID', 'Description'])
        
        if df is None:
            warning_msg = f"Could not load CSV for {sac_name}"
            logger.warning(warning_msg)
            self.warnings.append(warning_msg)
            print(f"  WARNING: Could not load CSV for {sac_name}")
            return pd.DataFrame(columns=['ID', 'Description'])
        
        # Check if ID column exists
        if 'ID' not in df.columns:
            # Try to find ID-like column
            id_cols = [c for c in df.columns if 'ID' in c.upper() or c == df.columns[0]]
            if id_cols:
                original_col = id_cols[0]
                df = df.rename(columns={original_col: 'ID'})
                logger.info(f"Using column '{original_col}' as ID for {dim_name}")
                print(f"  INFO: Using column '{original_col}' as ID")
            else:
                error_msg = f"No ID column found in {sac_name} (available: {list(df.columns)})"
                logger.error(error_msg)
                self.errors.append(error_msg)
                print(f"  WARNING: No ID column found in {sac_name}")
                return pd.DataFrame(columns=['ID', 'Description'])
        
        original_count = len(df)
        
        # Remove unassigned
        df = self.remove_unassigned(df)
        removed_unassigned = original_count - len(df)
        if removed_unassigned > 0:
            logger.debug(f"Removed {removed_unassigned} unassigned members from {dim_name}")
        
        # Apply filters
        pre_filter_count = len(df)
        df = self.apply_filters(df, filters)
        filtered_count = pre_filter_count - len(df)
        if filtered_count > 0:
            logger.debug(f"Filtered out {filtered_count} members from {dim_name}")
        
        # Get leaf members if hierarchy
        if has_hierarchy:
            parent_filter = filters.get("parent_filter")
            pre_leaf_count = len(df)
            df = self.get_leaf_members(df, parent_filter)
            leaf_filtered = pre_leaf_count - len(df)
            if leaf_filtered > 0:
                logger.debug(f"Filtered to leaf members: {leaf_filtered} parents removed from {dim_name}")
        
        # Apply ID list filter (after leaf filtering)
        id_list = filters.get("id_list", [])
        if id_list:
            pre_idlist_count = len(df)
            df = df[df['ID'].isin(id_list)].copy()
            if not df.empty:
                df['_sort'] = df['ID'].apply(lambda x: id_list.index(x) if x in id_list else 999)
                df = df.sort_values('_sort').drop('_sort', axis=1)
            idlist_filtered = pre_idlist_count - len(df)
            if idlist_filtered > 0:
                logger.debug(f"ID list filter removed {idlist_filtered} members from {dim_name}")
        
        # Return only ID and Description
        if 'Description' in df.columns:
            result = df[['ID', 'Description']].copy()
        else:
            result = df[['ID']].copy()
            result['Description'] = result['ID']
        
        result = result.dropna(subset=['ID'])
        
        if len(result) == 0:
            warning_msg = f"Dimension {dim_name} has 0 members after filtering"
            logger.warning(warning_msg)
            self.warnings.append(warning_msg)
        else:
            logger.info(f"Loaded dimension {dim_name}: {len(result)} members")
        
        return result
    
    def load_date_range(self):
        """Load date range from Version dimension or manual config."""
        # First check for manual date range in project config
        manual_range = self.project.get("date_range", {})
        if manual_range.get("start_month") and manual_range.get("end_month"):
            start_month = manual_range["start_month"]
            end_month = manual_range["end_month"]
            
            start_year, start_m = int(start_month[:4]), int(start_month[4:])
            end_year, end_m = int(end_month[:4]), int(end_month[4:])
            
            self.date_range = []
            year, month = start_year, start_m
            while (year, month) <= (end_year, end_m):
                self.date_range.append(f"{year}{month:02d}")
                month += 1
                if month > 12:
                    month = 1
                    year += 1
            print(f"  Using manual date range: {self.date_range[0]} - {self.date_range[-1]}")
            return
        
        # Try to load from VersionMaster.csv
        csv_path = os.path.join(self.downloads_dir, "VersionMaster.csv")
        if not os.path.exists(csv_path):
            print("  WARNING: VersionMaster.csv not found, using default date range")
            self.date_range = ["202501", "202502", "202503", "202504", "202505", "202506",
                              "202507", "202508", "202509", "202510", "202511", "202512"]
            return
        
        df = self.load_csv(csv_path)
        if df is None or 'ID' not in df.columns:
            print("  WARNING: Could not read Version CSV, using default date range")
            self.date_range = ["202501", "202502", "202503", "202504", "202505", "202506",
                              "202507", "202508", "202509", "202510", "202511", "202512"]
            return
            
        version_id = self.project.get("version", {}).get("version_id", "public.RF_CURRENT")
        
        version_row = df[df['ID'] == version_id]
        if version_row.empty:
            print(f"  WARNING: Version '{version_id}' not found, using default")
            self.date_range = ["202501", "202502", "202503", "202504", "202505", "202506",
                              "202507", "202508", "202509", "202510", "202511", "202512"]
            return
        
        try:
            # Get column names from config or use defaults
            version_config = self.project.get("version", {})
            start_col = version_config.get("start_column", "").strip() or "StartMonth"
            end_col = version_config.get("end_column", "").strip() or "EndMonth"
            
            # Check if columns exist
            if start_col not in df.columns or end_col not in df.columns:
                print(f"  WARNING: Columns '{start_col}' or '{end_col}' not found in Version CSV")
                print(f"  Available columns: {', '.join(df.columns.tolist())}")
                self.date_range = ["202501", "202502", "202503", "202504", "202505", "202506",
                                  "202507", "202508", "202509", "202510", "202511", "202512"]
                return
            
            start_month = str(version_row[start_col].values[0])
            end_month = str(version_row[end_col].values[0])
            
            # Generate month range
            start_year, start_m = int(start_month[:4]), int(start_month[4:])
            end_year, end_m = int(end_month[:4]), int(end_month[4:])
        
            self.date_range = []
            year, month = start_year, start_m
            while (year, month) <= (end_year, end_m):
                self.date_range.append(f"{year}{month:02d}")
                month += 1
                if month > 12:
                    month = 1
                    year += 1
            print(f"  Date range from Version: {self.date_range[0]} - {self.date_range[-1]}")
        except Exception as e:
            print(f"  WARNING: Error parsing Version dates: {e}, using default")
            self.date_range = ["202501", "202502", "202503", "202504", "202505", "202506",
                              "202507", "202508", "202509", "202510", "202511", "202512"]
    
    def create_excel(self, template: dict) -> OperationResult:
        """Create Excel template with comprehensive error handling."""
        template_name = template.get("name", "Unknown")
        logger.info(f"Creating Excel template: {template_name}")
        
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Get dimensions config as dict
            dimensions = {d["name"]: d for d in self.dimensions_config.get("dimensions", [])}
            
            # Apply template overrides
            overrides = template.get("dimension_overrides", {})
            for name, override in overrides.items():
                if name in dimensions:
                    dim_copy = dimensions[name].copy()
                    filters = dim_copy.get("filters", {}).copy()
                    filters.update(override)
                    dim_copy["filters"] = filters
                    dimensions[name] = dim_copy
                    logger.debug(f"Applied override for dimension: {name}")
                else:
                    warning_msg = f"Override specified for unknown dimension: {name}"
                    logger.warning(warning_msg)
                    self.warnings.append(warning_msg)
            
            columns = template.get("columns", [])
            
            if not columns:
                raise TemplateError(
                    f"Template '{template_name}' has no columns defined",
                    {"template": template_name}
                )
            
            # Verify all columns exist in dimensions
            missing_dims = [col for col in columns if col not in dimensions]
            if missing_dims:
                raise TemplateError(
                    f"Template '{template_name}' references undefined dimensions: {missing_dims}",
                    {"template": template_name, "missing": missing_dims}
                )
            
            # Load dimensions
            dimension_errors = []
            for col in columns:
                if col not in self.dimension_data:
                    dim_config = dimensions.get(col, {})
                    try:
                        self.dimension_data[col] = self.load_dimension(dim_config)
                        member_count = len(self.dimension_data[col])
                        print(f"    {col}: {member_count} members")
                        if member_count == 0:
                            dimension_errors.append(f"{col}: 0 members")
                    except DimensionError as e:
                        logger.error(f"Failed to load dimension {col}: {e}")
                        dimension_errors.append(f"{col}: {e.message}")
                        self.dimension_data[col] = pd.DataFrame(columns=['ID', 'Description'])
            
            if dimension_errors:
                logger.warning(f"Dimension loading issues for {template_name}: {dimension_errors}")
            
            # Create Upload sheet
            upload_sheet = wb.create_sheet("Upload_to_SAC")
            data_rows = template.get("data_rows", self.project.get("settings", {}).get("data_rows", 200))
            self._setup_upload_sheet(upload_sheet, columns, data_rows)
            
            # Create dimension sheets
            for col in columns:
                dim_config = dimensions.get(col, {})
                ws = wb.create_sheet(col[:31])  # Excel sheet name limit
                self._setup_dimension_sheet(ws, col, dim_config, self.dimension_data.get(col))
            
            # Add data validation
            self._add_data_validation(upload_sheet, columns, dimensions)
            
            # Save
            output_file = template.get("output_file", f"{template_name}.xlsx")
            output_path = os.path.join(self.output_dir, output_file)
            
            try:
                wb.save(output_path)
            except PermissionError:
                raise ExcelGenerationError(
                    f"Cannot save {output_file} - file may be open in Excel",
                    {"path": output_path}
                )
            except Exception as e:
                raise ExcelGenerationError(
                    f"Failed to save {output_file}: {str(e)}",
                    {"path": output_path}
                )
            
            logger.info(f"Successfully created: {output_path}")
            
            return OperationResult.ok(
                f"Created {output_file}",
                data={"path": output_path, "columns": columns, "dimension_errors": dimension_errors},
                warnings=dimension_errors if dimension_errors else None
            )
            
        except SACGeneratorError:
            raise
        except Exception as e:
            error_msg = f"Unexpected error creating template {template_name}: {str(e)}"
            logger.exception(error_msg)
            raise ExcelGenerationError(error_msg, {"template": template_name, "traceback": traceback.format_exc()})
    
    def _setup_upload_sheet(self, ws, columns: List[str], data_rows: int):
        """Setup Upload_to_SAC sheet."""
        headers = columns + self.date_range
        
        # Get colors from config
        colors = self.project.get("colors", {})
        dim_header_color = colors.get("dim_header", "#C6E0B4").replace("#", "")
        date_header_color = colors.get("date_header", "#BDD7EE").replace("#", "")
        dim_cell_color = colors.get("dim_cell", "#E2EFDA").replace("#", "")
        
        # Styles
        dim_header_fill = PatternFill('solid', fgColor=dim_header_color)
        date_header_fill = PatternFill('solid', fgColor=date_header_color)
        dim_cell_fill = PatternFill('solid', fgColor=dim_cell_color)
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        thick_right = Border(
            left=Side(style='thin'), right=Side(style='thick'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.fill = dim_header_fill if col_idx <= len(columns) else date_header_fill
        
        # Column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18 if col_idx <= len(columns) else 10
        
        # Data rows
        for row in range(2, data_rows + 2):
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = dim_cell_fill
                cell.border = thin_border
            for col_idx in range(len(columns) + 1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).border = thin_border
        
        # Thick border after dimensions
        for row in range(1, data_rows + 2):
            ws.cell(row=row, column=len(columns)).border = thick_right
    
    def _setup_dimension_sheet(self, ws, dim_name: str, dim_config: dict, data: pd.DataFrame):
        """Setup dimension sheet with table."""
        table_name = dim_config.get("table_name", f"tbl_{dim_name.lower().replace(' ', '_')}")
        
        # Title row
        title_fill = PatternFill('solid', fgColor='A9D08E')
        header_fill = PatternFill('solid', fgColor='C6E0B4')
        
        ws.merge_cells('A1:B1')
        ws['A1'] = dim_name
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        ws['A2'] = 'Description'
        ws['B2'] = 'ID'
        ws['A2'].font = Font(bold=True)
        ws['B2'].font = Font(bold=True)
        ws['A2'].fill = header_fill
        ws['B2'].fill = header_fill
        
        # Data
        if data is not None and not data.empty:
            for row_idx, row in enumerate(data.itertuples(), 3):
                ws.cell(row=row_idx, column=1, value=row.Description)
                ws.cell(row=row_idx, column=2, value=row.ID)
        
        # Column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        
        # Create table
        if data is not None and not data.empty:
            table_range = f"A2:B{len(data) + 2}"
            table = Table(displayName=table_name, ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(table)
    
    def _add_data_validation(self, ws, columns: List[str], dimensions: dict):
        """Add dropdown validation to Upload sheet."""
        data_rows = self.project.get("settings", {}).get("data_rows", 200)
        
        for col_idx, col_name in enumerate(columns, 1):
            dim_config = dimensions.get(col_name, {})
            table_name = dim_config.get("table_name", f"tbl_{col_name.lower().replace(' ', '_')}")
            
            dv = DataValidation(
                type="list",
                formula1=f'INDIRECT("{table_name}[ID]")',
                allowBlank=True
            )
            dv.error = f"Please select a valid {col_name}"
            dv.prompt = f"Select {col_name}"
            
            col_letter = get_column_letter(col_idx)
            dv.add(f"{col_letter}2:{col_letter}{data_rows + 1}")
            ws.add_data_validation(dv)
    
    def generate(self, template_name: Optional[str] = None) -> OperationResult:
        """Generate Excel templates with comprehensive error handling."""
        logger.info(f"Starting template generation (filter: {template_name or 'ALL'})")
        print("\n=== Generating Templates ===\n")
        
        # Reset errors/warnings for this run
        self.errors = []
        self.warnings = []
        
        templates = self.templates_config.get("templates", [])
        
        if not templates:
            error_msg = "No templates configured in config/templates.json"
            logger.error(error_msg)
            print(error_msg)
            return OperationResult.fail(error_msg)
        
        # Filter to specific template if requested
        if template_name:
            templates = [t for t in templates if t.get("name") == template_name]
            if not templates:
                error_msg = f"Template '{template_name}' not found"
                logger.error(error_msg)
                print(error_msg)
                available = [t.get("name") for t in self.templates_config.get("templates", [])]
                return OperationResult.fail(error_msg, errors=[f"Available templates: {available}"])
        
        # Load date range
        try:
            self.load_date_range()
            if self.date_range:
                print(f"Date range: {self.date_range[0]} - {self.date_range[-1]}\n")
            else:
                self.warnings.append("No date range loaded")
        except Exception as e:
            logger.exception("Failed to load date range")
            self.warnings.append(f"Date range error: {str(e)}")
        
        # Generate each template
        results = {"success": [], "failed": [], "warnings": []}
        
        for template in templates:
            name = template.get("name", "Unnamed")
            print(f"Generating: {name}")
            logger.info(f"Generating template: {name}")
            
            self.dimension_data = {}  # Reset for each template
            
            try:
                result = self.create_excel(template)
                if result.success:
                    results["success"].append({
                        "name": name,
                        "path": result.data.get("path"),
                        "warnings": result.warnings
                    })
                    results["warnings"].extend(result.warnings)
                    print(f"  ✓ Saved: {result.data.get('path')}\n")
                else:
                    results["failed"].append({"name": name, "errors": result.errors})
                    print(f"  ✗ Failed: {result.message}\n")
                    
            except SACGeneratorError as e:
                logger.error(f"Failed to generate {name}: {e}")
                results["failed"].append({"name": name, "error": e.message, "details": e.details})
                print(f"  ✗ Error: {e.message}\n")
                
            except Exception as e:
                logger.exception(f"Unexpected error generating {name}")
                results["failed"].append({"name": name, "error": str(e)})
                print(f"  ✗ Unexpected error: {str(e)}\n")
        
        # Summary
        success_count = len(results["success"])
        failed_count = len(results["failed"])
        
        logger.info(f"Generation complete: {success_count} success, {failed_count} failed")
        
        if failed_count > 0:
            print(f"\n⚠ Generated {success_count}, Failed {failed_count}")
            return OperationResult(
                status=ResultStatus.WARNING if success_count > 0 else ResultStatus.ERROR,
                message=f"Generated {success_count}, failed {failed_count}",
                data=results,
                errors=[f["error"] if "error" in f else str(f["errors"]) for f in results["failed"]],
                warnings=results["warnings"] + self.warnings
            )
        
        print(f"\n✓ Generated {success_count} template(s)")
        return OperationResult.ok(
            f"Generated {success_count} template(s)",
            data=results,
            warnings=results["warnings"] + self.warnings
        )


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="SAC Template Generator - Create Excel templates for SAP Analytics Cloud",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sac_generator.py --init              Initialize new project
  python sac_generator.py --download          Download dimension CSVs from SAC
  python sac_generator.py --scan              Scan CSVs and create dimensions config
  python sac_generator.py --validate          Validate configuration
  python sac_generator.py --generate          Generate all templates
  python sac_generator.py --generate "Name"   Generate specific template
  python sac_generator.py --log-level DEBUG   Enable debug logging
        """
    )
    
    parser.add_argument('--init', action='store_true', help='Initialize empty configuration')
    parser.add_argument('--download', action='store_true', help='Download dimension CSVs from SAC')
    parser.add_argument('--scan', action='store_true', help='Scan CSV files and suggest dimensions')
    parser.add_argument('--list', action='store_true', help='List configured templates')
    parser.add_argument('--validate', action='store_true', help='Validate configuration')
    parser.add_argument('--generate', nargs='?', const='ALL', help='Generate templates (optionally specify name)')
    parser.add_argument('--config', default='config', help='Configuration directory (default: config)')
    parser.add_argument('--log-level', default='INFO', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
                        help='Logging level (default: INFO)')
    
    args = parser.parse_args()
    
    # Setup logging with specified level
    global logger
    logger = setup_logging(level=args.log_level)
    
    # Show help if no arguments
    if len(sys.argv) == 1:
        parser.print_help()
        return
    
    try:
        generator = SACGenerator(config_dir=args.config)
        
        if args.init:
            generator.init_config()
        elif args.download:
            result = generator.download_all()
            if not result.success:
                sys.exit(1)
        elif args.scan:
            generator.scan_csv_files()
        elif args.list:
            generator.list_templates()
        elif args.validate:
            generator.validate_config()
        elif args.generate:
            template_name = None if args.generate == 'ALL' else args.generate
            result = generator.generate(template_name)
            if not result.success:
                sys.exit(1)
                
    except SACGeneratorError as e:
        logger.error(f"Error: {e.message}")
        print(f"\n❌ Error: {e.message}")
        if e.details:
            print(f"   Details: {e.details}")
        sys.exit(1)
        
    except KeyboardInterrupt:
        logger.info("Operation cancelled by user")
        print("\n\nOperation cancelled.")
        sys.exit(130)
        
    except Exception as e:
        logger.exception("Unexpected error")
        print(f"\n❌ Unexpected error: {str(e)}")
        print("   Check logs/sac_generator_*.log for details")
        sys.exit(1)


if __name__ == "__main__":
    main()
